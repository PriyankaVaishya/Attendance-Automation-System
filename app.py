import numpy as np
import cv2
import firebase_admin
from firebase_admin import credentials,firestore
# from firebase_admin import storage
import pyrebase
import urllib
import urlopen
import keras
import math
from firebase import Firebase
from google.oauth2 import service_account
import openpyxl
from datetime import date
from PIL import Image
import firebase
from scipy import ndimage
import argparse
import imutils
import orientation1
import segmentation
import os
from google.cloud import storage
import google.auth.transport.requests
from google.auth.transport.requests import AuthorizedSession

def getBestShift(img):
    cy,cx = ndimage.measurements.center_of_mass(img)

    rows,cols = img.shape
    shiftx = np.round(cols/2.0-cx).astype(int)
    shifty = np.round(rows/2.0-cy).astype(int)

    return shiftx,shifty
def shift(img,sx,sy):
    rows,cols = img.shape
    M = np.float32([[1,0,sx],[0,1,sy]])
    shifted = cv2.warpAffine(img,M,(cols,rows))
    return shifted

def preprocess(x):
    
    while np.sum(x[0]) == 0:
        x = x[1:]

    while np.sum(x[:,0]) == 0:
        x = np.delete(x,0,1)

    while np.sum(x[-1]) == 0:
        x = x[:-1]

    while np.sum(x[:,-1]) == 0:
        x = np.delete(x,-1,1)

    rows,cols = x.shape
    if rows > cols:
        factor = 20.0/rows
        rows = 20
        cols = int(round(cols*factor))
        x = cv2.resize(x, (cols,rows))
    else:
        factor = 20.0/cols
        cols = 20
        rows = int(round(rows*factor))
        x = cv2.resize(x, (cols, rows))
    colsPadding = (int(math.ceil((28-cols)/2.0)),int(math.floor((28-cols)/2.0)))
    rowsPadding = (int(math.ceil((28-rows)/2.0)),int(math.floor((28-rows)/2.0)))
    x = np.lib.pad(x,(rowsPadding,colsPadding),'constant')
    shiftx,shifty = getBestShift(x)
    shifted = shift(x,shiftx,shifty)
    x = shifted
    return x

def order_points(pts):
	# initialzie a list of coordinates that will be ordered
	# such that the first entry in the list is the top-left,
	# the second entry is the top-right, the third is the
	# bottom-right, and the fourth is the bottom-left
	rect = np.zeros((4, 2), dtype = "float32")
 
	# the top-left point will have the smallest sum, whereas
	# the bottom-right point will have the largest sum
	s = pts.sum(axis = 1)
	rect[0] = pts[np.argmin(s)]
	rect[2] = pts[np.argmax(s)]
 
	# now, compute the difference between the points, the
	# top-right point will have the smallest difference,
	# whereas the bottom-left will have the largest difference
	diff = np.diff(pts, axis = 1)
	rect[1] = pts[np.argmin(diff)]
	rect[3] = pts[np.argmax(diff)]
 
	# return the ordered coordinates
	return rect

def four_point_transform(image, pts):
	# obtain a consistent order of the points and unpack them
	# individually
	rect = order_points(pts)
	(tl, tr, br, bl) = rect
 
	# compute the width of the new image, which will be the
	# maximum distance between bottom-right and bottom-left
	# x-coordiates or the top-right and top-left x-coordinates
	widthA = np.sqrt(((br[0] - bl[0]) ** 2) + ((br[1] - bl[1]) ** 2))
	widthB = np.sqrt(((tr[0] - tl[0]) ** 2) + ((tr[1] - tl[1]) ** 2))
	maxWidth = max(int(widthA), int(widthB))
 
	# compute the height of the new image, which will be the
	# maximum distance between the top-right and bottom-right
	# y-coordinates or the top-left and bottom-left y-coordinates
	heightA = np.sqrt(((tr[0] - br[0]) ** 2) + ((tr[1] - br[1]) ** 2))
	heightB = np.sqrt(((tl[0] - bl[0]) ** 2) + ((tl[1] - bl[1]) ** 2))
	maxHeight = max(int(heightA), int(heightB))
 
	# now that we have the dimensions of the new image, construct
	# the set of destination points to obtain a "birds eye view",
	# (i.e. top-down view) of the image, again specifying points
	# in the top-left, top-right, bottom-right, and bottom-left
	# order
	dst = np.array([
		[0, 0],
		[maxWidth - 1, 0],
		[maxWidth - 1, maxHeight - 1],
		[0, maxHeight - 1]], dtype = "float32")
 
	# compute the perspective transform matrix and then apply it
	M = cv2.getPerspectiveTransform(rect, dst)
	warped = cv2.warpPerspective(image, M, (maxWidth, maxHeight))
 
	# return the warped image
	return warped
def image_resize(image, width = None, height = None, inter = cv2.INTER_AREA):
    # initialize the dimensions of the image to be resized and
    # grab the image size
    dim = None
    (h, w) = image.shape[:2]

    # if both the width and height are None, then return the
    # original image
    if width is None and height is None:
        return image

    # check to see if the width is None
    if width is None:
        # calculate the ratio of the height and construct the
        # dimensions
        r = height / float(h)
        dim = (int(w * r), height)

    # otherwise, the height is None
    else:
        # calculate the ratio of the width and construct the
        # dimensions
        r = width / float(w)
        dim = (width, int(h * r))

    # resize the image
    resized = cv2.resize(image, dim, interpolation = inter)

    # return the resized image
    return resized
def get_contour_precedence(contour, cols):
    origin = cv2.boundingRect(contour)
    return origin[1] * cols + origin[0]



config= {
	"apiKey": "AIzaSyDYHyTfB4u_OswwN507ql9kXx5sdedXZEg",
	"project_id": "attendance-app-1c683",
	"authDomain": "attendance-app-1c683.firebaseapp.com",
	"storageBucket": "attendance-app-1c683.appspot.com",
	"databaseURL" : "https://attendance-app-1c683.firebaseio.com",
	"messagingSenderId": "522004839232"
}

def main():
	firebase =Firebase(config)
	db=firebase.database()
	while(True):
		users=db.child("uploads").child("abc").child("a").get()
		if users.val()!=None:
			urllib.request.urlretrieve(str(users.val()),"test.jpg")
			break
	db.child("uploads").remove()

	new_model = keras.models.load_model('cnn.h5')

	try:
		# read image
		img = cv2.imread('test.jpg')
		img=image_resize(img,width=800,height=700)
		image=orientation1.angle(img)
		gray=image
		ret,thresh = cv2.threshold(gray,127,255,cv2.THRESH_BINARY_INV)
		kernel = np.ones((5,100), np.uint8)
		img_dilation = cv2.dilate(thresh, kernel, iterations=1)
		# cv2.imshow('dilated',img_dilation)
		# cv2.waitKey(0)
		# cv2.imwrite('dilated.jpg', img_dilation) 
		#find contours
		_,ctrs, hier = cv2.findContours(img_dilation.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
		#sort contours
		sorted_ctrs = sorted(ctrs, key=lambda ctr: (cv2.boundingRect(ctr)[0],cv2.boundingRect(ctr)[1]))
		output=[]
		for i, ctr in enumerate(sorted_ctrs):
			# Get bounding box
			x, y, w, h = cv2.boundingRect(ctr)

			# Getting ROI
			roi = image[y:y+h, x:x+w]
			area=w*h

			if h<w and h> 10 and 100 < w < 500 and area < 12000 :
				grayy= roi
				# roi=image_resize(roi,width=100,height=30)
				# cv2.rectangle(img,(x,y),(x+w,y+h),(0,255,0),1)
				ret,threshh = cv2.threshold(grayy,127,255,cv2.THRESH_BINARY_INV)
				# cv2.imshow('inner',threshh)
				# cv2.waitKey(0)
				# cv2.imwrite('word.jpg', threshh) 
				_,ctrss, hierr = cv2.findContours(threshh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
				sorted_ctrss = sorted(ctrss, key=lambda ctr: cv2.boundingRect(ctr)[0])
				num=0
				for ii, ctrr in enumerate(sorted_ctrss):
					# Get bounding box
					xx, yy, ww, hh = cv2.boundingRect(ctrr)

					# Getting ROI
					roii = roi[yy-2:yy+hh+2, xx-1:xx+ww+1]
					if  1 < ww and 4 < hh :
						# cv2.imshow('wer ',roii)
						# cv2.waitKey(0)
						# cv2.rectangle(img,(xx,yy),(xx+ww,yy+hh),(0,255,0),1)

						# cv2.imwrite('single.jpg', roii) 
						pic=image_resize(roii,width=28,height=28)

						(thresh, pic) = cv2.threshold(pic, 120, 255, cv2.THRESH_BINARY_INV)
						pic=preprocess(pic)
						# cv2.imshow('inner',pic)
						# cv2.waitKey(0)
						# cv2.imwrite('singlemnist.jpg', pic) 
						pic=np.array(pic).reshape(28,28,1)
						pic=np.expand_dims(pic,axis=0)                
						prediction=new_model.predict(pic)
						x=np.argmax(prediction)
						num=num*10+x
				output.append(num)
		output.sort()
		file=open("output.txt","w+")
		for i in output:
			file.write(str(i))
			file.write('\n')
		file.close()
		#entering attendence in sheet
		# cv2.imwrite('final.jpg',img)
		loc=("/home/pankaj/Desktop/minor/attendence.xlsx")
		wb1 = openpyxl.load_workbook(loc) 
		ws1 = wb1.worksheets[0] 
		wb2=openpyxl.Workbook()
		ws2=wb2.active
		nrow=ws1.max_row
		ncol=ws1.max_column
		for i in range(1,200):
			c1=ws1.cell(row=i,column=1)
			if not c1.value:
				nrow=i
				break
		for i in range(1,200):
			c1=ws1.cell(row=1,column=i)
			if not c1.value:
				ncol=i
				break
		print(ncol)
		for i in range(1,nrow):
			for j in range(1,ncol):
				copyfrom=ws1.cell(row=i,column=j)
				ws2.cell(row=i,column=j).value=copyfrom.value
		ncol=ncol
		today=date.today()
		today=today.strftime("%m/%d/%Y")
		c1=ws2.cell(row=1,column=ncol)
		c1.value=today
		for i in range(2,nrow+1):
			c1=ws2.cell(row=i,column=2)
			# flag=0
			for j in range(len(output)):
				if c1.value == output[j]:
					ws2.cell(row=i,column=ncol).value="P"

		for i in range(2,nrow):
			c1=ws2.cell(row=i,column=ncol)
			if not c1.value:
				c1.value="A"
		#if attendence if of same day reject
		if ws2.cell(row=1,column=ncol).value == ws2.cell(row=1,column=ncol-1).value:
			ws2.delete_cols(ncol)
		wb2.save("/home/pankaj/Desktop/minor/attendence.xlsx")
	except:
		print("error occured")
		return
	storage=firebase.storage()
	storage.child("attendence.xlsx").put("attendence.xlsx")

	return
if __name__ == '__main__':
	firebase =Firebase(config)
	db=firebase.database()
	db.child("uploads").remove()
	while(True):
		main()